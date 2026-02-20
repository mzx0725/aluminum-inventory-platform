"""
入库管理路由蓝图
"""
from io import BytesIO
from flask import Blueprint, request, jsonify, send_file
from sqlalchemy import and_, or_
from app import db
from app.models import InboundRecord, Inventory, AluminumPlate, User
from app.utils.time_utils import get_beijing_time

inbound_bp = Blueprint('inbound', __name__)


@inbound_bp.route('', methods=['POST'])
def create_inbound():
    """创建入库记录（入库登记）"""
    data = request.get_json()
    
    if not data or not data.get('plate_id') or not data.get('quantity'):
        return jsonify({'error': '铝板ID和入库数量不能为空'}), 400
    
    plate_id = data['plate_id']
    quantity = data['quantity']
    operator_name = data.get('operator_name', '未知操作员')
    batch_number = data.get('batch_number')
    supplier = data.get('supplier')
    remark = data.get('remark')
    
    if quantity <= 0:
        return jsonify({'error': '入库数量必须大于0'}), 400
    
    plate = AluminumPlate.query.get(plate_id)
    if not plate:
        return jsonify({'error': '铝板不存在'}), 404
    
    operator = User.query.filter(
        or_(
            User.real_name == operator_name,
            User.username == operator_name
        )
    ).first()
    
    if not operator:
        operator = User(
            username=operator_name.lower().replace(' ', '_'),
            real_name=operator_name,
            role='warehouse',
            status='active'
        )
        operator.set_password('123456')
        db.session.add(operator)
        db.session.flush()
    
    try:
        inbound_record = InboundRecord(
            plate_id=plate_id,
            quantity=quantity,
            batch_number=batch_number,
            supplier=supplier,
            operator_id=operator.id,
            inbound_time=get_beijing_time(),
            remark=remark
        )
        
        db.session.add(inbound_record)
        
        inventory = None
        if batch_number:
            inventory = Inventory.query.filter_by(
                plate_id=plate_id,
                batch_number=batch_number
            ).first()
        
        if inventory:
            inventory.quantity += quantity
            inventory.last_updated = get_beijing_time()
        else:
            inventory = Inventory(
                plate_id=plate_id,
                quantity=quantity,
                batch_number=batch_number,
                location=data.get('location'),
                warning_threshold=data.get('warning_threshold', 10)
            )
            db.session.add(inventory)
        
        db.session.commit()
        
        result = inbound_record.to_dict()
        result['operator'] = operator_name
        
        return jsonify({
            'message': '入库登记成功',
            'inbound_record': result,
            'inventory': inventory.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'入库登记失败: {str(e)}'}), 500


@inbound_bp.route('', methods=['GET'])
def get_inbound_list():
    """获取入库记录列表（支持分页、时间范围筛选、铝板型号筛选）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    plate_model = request.args.get('plate_model')
    batch_number = request.args.get('batch_number')
    supplier = request.args.get('supplier')
    
    query = InboundRecord.query
    
    if start_date:
        try:
            start_datetime = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(InboundRecord.inbound_time >= start_datetime)
        except ValueError:
            return jsonify({'error': '开始时间格式错误'}), 400
    
    if end_date:
        try:
            end_datetime = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(InboundRecord.inbound_time <= end_datetime)
        except ValueError:
            return jsonify({'error': '结束时间格式错误'}), 400
    
    if plate_model:
        query = query.join(AluminumPlate).filter(
            AluminumPlate.model.ilike(f'%{plate_model}%')
        )
    
    if batch_number:
        query = query.filter(InboundRecord.batch_number.ilike(f'%{batch_number}%'))
    
    if supplier:
        query = query.filter(InboundRecord.supplier.ilike(f'%{supplier}%'))
    
    query = query.order_by(InboundRecord.inbound_time.desc())
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    records = []
    for record in pagination.items:
        record_dict = record.to_dict()
        if record.plate:
            record_dict['plate'] = record.plate.to_dict()
        if record.operator:
            record_dict['operator_name'] = record.operator.real_name
        records.append(record_dict)
    
    return jsonify({
        'list': records,
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages
    }), 200


@inbound_bp.route('/<int:inbound_id>', methods=['GET'])
def get_inbound_detail(inbound_id):
    """获取入库记录详情"""
    inbound_record = InboundRecord.query.get(inbound_id)
    
    if not inbound_record:
        return jsonify({'error': '入库记录不存在'}), 404
    
    return jsonify({
        'inbound_record': inbound_record.to_dict()
    }), 200


@inbound_bp.route('/export', methods=['GET'])
def export_inbound():
    """导出入库记录为Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return jsonify({'error': '缺少openpyxl库，请先安装: pip install openpyxl'}), 500
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    plate_model = request.args.get('plate_model')
    
    query = InboundRecord.query
    
    if start_date:
        try:
            start_datetime = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(InboundRecord.inbound_time >= start_datetime)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_datetime = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(InboundRecord.inbound_time <= end_datetime)
        except ValueError:
            pass
    
    if plate_model:
        query = query.join(AluminumPlate).filter(
            AluminumPlate.model.ilike(f'%{plate_model}%')
        )
    
    records = query.order_by(InboundRecord.inbound_time.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '入库记录'
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['序号', '铝板型号', '铝板规格', '入库数量', '单位', '批次号', 
               '供应商', '操作员', '入库时间', '备注']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_num, record in enumerate(records, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=record.plate.model if record.plate else '').border = thin_border
        ws.cell(row=row_num, column=3, value=record.plate.specification if record.plate else '').border = thin_border
        ws.cell(row=row_num, column=4, value=record.quantity).border = thin_border
        ws.cell(row=row_num, column=5, value=record.plate.unit if record.plate else '').border = thin_border
        ws.cell(row=row_num, column=6, value=record.batch_number or '').border = thin_border
        ws.cell(row=row_num, column=7, value=record.supplier or '').border = thin_border
        ws.cell(row=row_num, column=8, value=record.operator.real_name if record.operator else '').border = thin_border
        
        inbound_time = record.inbound_time.strftime('%Y-%m-%d %H:%M:%S') if record.inbound_time else ''
        ws.cell(row=row_num, column=9, value=inbound_time).border = thin_border
        ws.cell(row=row_num, column=10, value=record.remark or '').border = thin_border
    
    column_widths = [8, 15, 20, 12, 8, 15, 20, 12, 20, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'入库记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )